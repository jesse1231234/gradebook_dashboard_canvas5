import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import re, io
from datetime import timedelta

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import DataBarRule, CellIsRule
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Gradebook + Echo Analyzer", page_icon="🎓", layout="wide")

# ---------------- Utilities shared ----------------
def shorten_label(s, limit=42):
    s2 = re.sub(r"\(\d+\)$", "", str(s)).strip()
    return (s2[:limit] + "…") if len(s2) > limit else s2

# ---------------- Echo helpers (Echo Formatting 5.py parity) ----------------
def time_to_seconds(ts: str) -> int:
    if pd.isna(ts) or ts == "":
        return 0
    parts = list(map(int, str(ts).split(":")))
    while len(parts) < 3:
        parts.insert(0, 0)
    h, m, s = parts
    return h*3600 + m*60 + s

def seconds_to_hms(sec: float) -> str:
    return "" if pd.isna(sec) else str(timedelta(seconds=int(sec)))

def natural_key(s: str):
    return [int(chunk) if chunk.isdigit() else chunk.lower() for chunk in re.split(r'(\d+)', str(s))]

def echo_analyze(df: pd.DataFrame) -> pd.DataFrame:
    # 1) parse
    df = df.copy()
    df['Duration_sec']      = df['Duration'].apply(time_to_seconds)
    df['TotalViewTime_sec'] = df['Total View Time'].apply(time_to_seconds)
    df['AvgViewTime_sec']   = df['Average View Time'].apply(time_to_seconds)
    df['Row View %']        = df['TotalViewTime_sec'] / df['Duration_sec'].replace(0, np.nan)

    # 2) group
    grp = df.groupby('Media Name', sort=False)
    titles = list(grp.groups.keys())
    media_count = len(titles)

    # 3) core summary
    summary_core = pd.DataFrame({
        'Media Title':              titles,
        'Video Duration':           [grp.get_group(t)['Duration_sec'].iloc[0] for t in titles],
        'Number of Unique Viewers': grp['User Name'].nunique().values,
        'Average View %':           grp['Row View %'].mean().fillna(0).values,
        'Total View %':             (grp['TotalViewTime_sec'].sum() / grp['Duration_sec'].sum()).values,
        'Total View Time':          grp['TotalViewTime_sec'].sum().values,
        'Average View Time':        grp['AvgViewTime_sec'].mean().values,
        'Average Total View Time':  grp['TotalViewTime_sec'].mean().values,
    })

    # 4) natural sort
    summary_core['sort_key'] = summary_core['Media Title'].apply(natural_key)
    summary_core = summary_core.sort_values('sort_key').drop(columns='sort_key').reset_index(drop=True)

    # 5) Grand Total (averages)
    means = summary_core[['Video Duration','Total View Time','Average View Time','Average Total View Time']].mean()
    viewers_mean = summary_core['Number of Unique Viewers'].mean()
    summary_core.loc[len(summary_core)] = {
        'Media Title':               'Grand Total',
        'Video Duration':            means['Video Duration'],
        'Number of Unique Viewers':  viewers_mean,
        'Average View %':            summary_core['Average View %'].mean(),
        'Total View %':              summary_core['Total View %'].mean(),
        'Total View Time':           means['Total View Time'],
        'Average View Time':         means['Average View Time'],
        'Average Total View Time':   means['Average Total View Time'],
    }

    # 6) Average Video Length and Watch Time
    n = len(summary_core) - 1
    means2 = summary_core.loc[:n-1, ['Video Duration','Total View Time','Average View Time','Average Total View Time']].mean()
    summary_core.loc[len(summary_core)] = {
        'Media Title':               'Average Video Length and Watch Time',
        'Video Duration':            means2['Video Duration'],
        'Number of Unique Viewers':  '',
        'Average View %':            summary_core.loc[:n-1, 'Average View %'].mean(),
        'Total View %':              summary_core.loc[:n-1, 'Total View %'].mean(),
        'Total View Time':           means2['Total View Time'],
        'Average View Time':         means2['Average View Time'],
        'Average Total View Time':   means2['Average Total View Time'],
    }
    return summary_core

def echo_build_workbook(summary_df: pd.DataFrame) -> bytes:
    # Write like the script: with formatting, data bars, charts
    wb = Workbook()
    ws = wb.active
    ws.title = 'Media Summary'

    # Before converting, store seconds as Excel time for time columns B,F,G,H later
    # But initial write uses strings; then convert to Excel time numeric (days)
    tmp = summary_df.copy()
    # Convert seconds->string for display, like script
    for col in ['Video Duration','Total View Time','Average View Time','Average Total View Time']:
        tmp[col] = tmp[col].apply(seconds_to_hms)

    for row in dataframe_to_rows(tmp, index=False, header=True):
        ws.append(row)
    last_row = ws.max_row
    media_count = len(summary_df) - 2  # exclude the two footer rows

    # Convert Video Duration (col B) back to numeric Excel time
    for r in range(2, last_row + 1):
        cell = ws[f'B{r}']
        secs = time_to_seconds(cell.value)
        cell.value = secs / 86400.0
        cell.number_format = 'hh:mm:ss'

    # Percent format and time formats
    for r in range(2, last_row + 1):
        for col in ('D','E'):
            c = ws[f'{col}{r}']
            if isinstance(c.value, (int, float)):
                c.number_format = '0.00%'
        for col in ('F','G','H'):
            ws[f'{col}{r}'].number_format = 'hh:mm:ss'

    # Data bars (orange) for B and D on the top 'media_count' rows only
    if media_count >= 1:
        bar = DataBarRule(start_type='min', end_type='max', color="FFA500")
        ws.conditional_formatting.add(f"B2:B{1+media_count}", bar)
        ws.conditional_formatting.add(f"D2:D{1+media_count}", bar)

    # Charts
    chart1 = LineChart()
    chart1.title = "View % Over Time"
    chart1.style = 9
    chart1.y_axis.number_format = '0.00%'
    data1 = Reference(ws, min_col=4, min_row=1, max_row=1+media_count)
    chart1.add_data(data1, titles_from_data=True)
    cats = Reference(ws, min_col=1, min_row=2, max_row=1+media_count)
    chart1.set_categories(cats)
    ws.add_chart(chart1, "J2")

    chart2 = LineChart()
    chart2.title = "Unique Viewers Over Time"
    chart2.style = 9
    data2 = Reference(ws, min_col=3, min_row=1, max_row=1+media_count)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    ws.add_chart(chart2, "J20")

    # Table styling
    tbl = Table(displayName="MediaStats", ref=f"A1:H{last_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tbl.tableStyleInfo = style
    ws.add_table(tbl)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ---------------- Gradebook helpers (Gradebook Formatting 15 GOOD.py parity) ----------------
def gradebook_process(df: pd.DataFrame) -> pd.DataFrame:
    # Steps mirror the script (no Excel-specific parts here; those are in workbook builder)
    df = df.copy()

    # 2. Drop any row where the first column contains "Student, Test"
    mask = df.iloc[:, 0].astype(str).str.contains("Student, Test", na=False)
    df = df[~mask].reset_index(drop=True)

    # 3. Drop unwanted columns (NOT "Final Grade")
    to_drop = ["Student","ID","SIS User ID","SIS Login ID","Current Grade","Unposted Current Grade","Unposted Final Grade"]
    df.drop(columns=[c for c in to_drop if c in df.columns], inplace=True, errors="ignore")

    # 4. Drop any column where rows 3+ are all empty or only zeros, except "Final Grade"
    drop_cols = []
    for col in df.columns:
        if col == "Final Grade":
            continue
        s = pd.to_numeric(df[col].iloc[2:], errors='coerce')
        if s.fillna(0).eq(0).all():
            drop_cols.append(col)
    df.drop(columns=drop_cols, inplace=True, errors="ignore")
    return df

def gradebook_build_workbook(df: pd.DataFrame) -> bytes:
    # Reproduce the Excel-writing logic including formulas and formatting
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active

    # Write DataFrame directly
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # Find the column index of "Final Grade"
    final_grade_idx_pre = None
    for ci in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=ci).value == "Final Grade":
            final_grade_idx_pre = ci
            break

    # Fill empties with 0 in cols 2+, skipping Final Grade
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        for cell in row:
            if cell.column == final_grade_idx_pre:
                continue
            if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):
                cell.value = 0

    # Convert strings->numbers where possible, skipping Final Grade
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.column == final_grade_idx_pre:
                continue
            if isinstance(cell.value, str) and cell.value.strip():
                txt = cell.value.replace(",", "")
                try:
                    cell.value = float(txt)
                except ValueError:
                    pass

    # Replace "(read only)" in row 2 with column max of rows 3+ (skip Final Grade)
    data_last_row = ws.max_row
    for ci in range(1, ws.max_column + 1):
        if ci == final_grade_idx_pre:
            continue
        hdr = ws.cell(row=2, column=ci)
        if isinstance(hdr.value, str) and "(read only)" in hdr.value:
            nums = [ws.cell(row=r, column=ci).value for r in range(3, data_last_row + 1) if isinstance(ws.cell(row=r, column=ci).value, (int, float))]
            if nums:
                hdr.value = max(nums)

    # Insert new blank column A for Row Titles
    ws.insert_cols(1)
    ws["A1"] = "Row Titles"

    # Re-find "Final Grade" now that columns shifted
    final_grade_idx = None
    for ci in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=ci).value == "Final Grade":
            final_grade_idx = ci
            break

    # Label row 2 as Points Possible
    ws["A2"] = "Points Possible"

    # Append Average and Average Excluding Zeros rows
    original_last_data_row = ws.max_row
    avg_row  = original_last_data_row + 1
    avg0_row = original_last_data_row + 2
    ws[f"A{avg_row}"]  = "Average"
    ws[f"A{avg0_row}"] = "Average Excluding Zeros"

    # Fill in Average formulas & % formatting (skip Final Grade)
    max_col = ws.max_column
    for col in range(2, max_col + 1):
        if col == final_grade_idx:
            continue
        letter   = get_column_letter(col)
        data_rng = f"{letter}3:{letter}{original_last_data_row}"
        header   = f"{letter}$2"
        c_avg    = ws[f"{letter}{avg_row}"]
        c_avg.value = f"=AVERAGE({data_rng})/{header}"
        c_avg.number_format = '0.00%'
        c_avg0   = ws[f"{letter}{avg0_row}"]
        c_avg0.value = f"=AVERAGEIF({data_rng},\">0\")/{header}"
        c_avg0.number_format = '0.00%'

    # Conditional formatting on averages
    green  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    red    = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    for row in (avg_row, avg0_row):
        rng = f"B{row}:{get_column_letter(max_col)}{row}"
        ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThan', formula=['0.9'], fill=green))
        ws.conditional_formatting.add(rng, CellIsRule(operator='between',     formula=['0.8','0.9'], fill=yellow))
        ws.conditional_formatting.add(rng, CellIsRule(operator='lessThan',    formula=['0.8'], fill=red))

    # Count of F and Percent of F in Final Grade
    count_row = avg0_row + 1
    pct_row   = avg0_row + 2
    ws[f"A{count_row}"] = "Count of F"
    ws[f"A{pct_row}"]   = "Percent of F"
    fg_letter = get_column_letter(final_grade_idx)
    ws.cell(row=count_row, column=final_grade_idx).value = f'=COUNTIF({fg_letter}3:{fg_letter}{original_last_data_row},"F")'
    total_students = original_last_data_row - 2
    ws.cell(row=pct_row, column=final_grade_idx).value = f'={fg_letter}{count_row}/{total_students}'
    ws.cell(row=pct_row, column=final_grade_idx).number_format = '0.00%'

    # Format the data region as an Excel Table
    table_end = get_column_letter(max_col) + str(original_last_data_row)
    table = Table(displayName="GradesTable", ref=f"A1:{table_end}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ---------------- UI ----------------
st.title("🎓 Script-Accurate Analyzer")
st.caption("Mirrors the logic of your two Python scripts and lets you download Excel files with the same formulas and formatting.")

tabA, tabB = st.tabs(["📘 Gradebook (script parity)", "🎬 Echo (script parity)"])

# ===== Gradebook tab =====
with tabA:
    st.subheader("Upload a Gradebook CSV")
    gb = st.file_uploader("Choose a gradebook CSV", type=["csv"], key="gb_up")
    if not gb:
        st.info("Or try the sample: `samples/sample_grades.csv` in the repo.")
    else:
        df_raw = pd.read_csv(gb)
        df_proc = gradebook_process(df_raw)

        # Preview (identifying columns already removed per script)
        st.write("**Processed preview (identifiers removed; empty/zero-only assignment columns dropped):**")
        st.dataframe(df_proc.head(20), use_container_width=True)

        # Build workbook bytes
        try:
            xbytes = gradebook_build_workbook(df_proc)
            st.download_button("⬇️ Download Excel (script formulas & formatting)",
                               data=xbytes, file_name="Gradebook_Analyzed.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            st.error(f"Excel build failed: {e}")

# ===== Echo tab =====
with tabB:
    st.subheader("Upload an Echo CSV")
    ec = st.file_uploader("Choose an Echo CSV", type=["csv"], key="echo_up")
    if not ec:
        st.info("Or try the sample: `samples/sample_echo.csv` in the repo.")
    else:
        df = pd.read_csv(ec, dtype=str)
        # Ensure required columns exist
        missing_cols = [c for c in ["Media Name","Duration","User Name","Total View Time","Average View Time"] if c not in df.columns]
        if missing_cols:
            st.error(f"Missing required columns: {missing_cols}")
        else:
            summary = echo_analyze(df)

            # Display as a nice table (time columns human-readable)
            disp = summary.copy()
            for col in ['Video Duration','Total View Time','Average View Time','Average Total View Time']:
                disp[col] = disp[col].apply(lambda x: seconds_to_hms(x) if isinstance(x, (int,float,np.integer,np.floating)) else x)
            st.write("**Media Summary**")
            st.dataframe(disp, use_container_width=True)

            # Charts mirroring script intent
            # Average View % Over Time
            media_count = max(0, len(summary) - 2)
            if media_count > 0:
                main_rows = summary.iloc[:media_count]
                fig1 = go.Figure()
                fig1.add_trace(go.Scatter(x=main_rows['Media Title'], y=main_rows['Average View %'], mode='lines+markers', name='Average View %'))
                fig1.update_yaxes(tickformat=".0%")
                fig1.update_layout(title="View % Over Time")
                st.plotly_chart(fig1, use_container_width=True)

                # Unique viewers over time
                fig2 = go.Figure()
                fig2.add_trace(go.Scatter(x=main_rows['Media Title'], y=main_rows['Number of Unique Viewers'], mode='lines+markers', name='Unique Viewers'))
                fig2.update_layout(title="Unique Viewers Over Time")
                st.plotly_chart(fig2, use_container_width=True)

            # Excel download with charts/formatting
            try:
                xbytes = echo_build_workbook(summary)
                st.download_button("⬇️ Download Excel (script formatting & charts)",
                                   data=xbytes, file_name="Echo_Analyzed.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.error(f"Excel build failed: {e}")